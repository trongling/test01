from openpyxl import load_workbook
from logging_method2 import LoggingMethod

class ExcelMethod():
	def __init__(self,logger,filename,sheetName):
		self.filename=filename
		self.wb = load_workbook(filename)
		# 通过工作表名获取一个工作表对象
		self.sheet = self.wb.get_sheet_by_name(sheetName)
		# 获取工作表中的最大行号
		# self.maxRowNum=self.sheet.max_row
		# 获取工作表中的最大列号
		self.max_column = self.sheet.max_column
		self.logger = logger

	def readExcel(self):
		dataList=[]
		try:
			for row in self.sheet:
				tmpList = []
				for cell in row:
					tmpList.append(cell.value)
				dataList.append(tmpList)
		except:
			self.logger.error('%s加载失败' %self.filename)
		else:
			return dataList

	def saveExcel(self,row,text):
		try:
			self.sheet.cell(row,self.max_column,text)
			self.wb.save(self.filename)
		except:
			self.logger.error('%s 保存失败' %self.filename)

if __name__=="__main__":

	logger = LoggingMethod().getlogger()
	# excel = ExcelMethod(logger, "test.xlsx", "Sheet1")
	# excel.readExcel()
	data_list = ExcelMethod(logger, "test.xlsx", "Sheet1")
	cell_list= data_list.readExcel()
	print(cell_list)

	# for i in cell_list():
	# 	print(i[0],i[1],i[2])
	# 	print(i)